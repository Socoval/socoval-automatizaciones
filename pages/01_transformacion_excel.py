import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, Alignment

st.set_page_config(
    page_title="Control Facturas | SOCOVAL",
    page_icon="📊",
    layout="centered",
    initial_sidebar_state="collapsed"
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600&display=swap');
html, body, [data-testid="stAppViewContainer"], [data-testid="stMain"], .main, .block-container {
    background-color: #F4F6F9 !important;
    font-family: 'DM Sans', 'Segoe UI', sans-serif !important;
}
#MainMenu, footer, header { visibility: hidden !important; }
[data-testid="stSidebar"] { display: none !important; }
[data-testid="stToolbar"] { display: none !important; }
[data-testid="stDecoration"] { display: none !important; }
[data-testid="stHeader"] { display: none !important; }
.stDeployButton { display: none !important; }
.block-container { padding: 2rem 2rem !important; max-width: 780px !important; }

.firma {
    position: fixed;
    bottom: 14px;
    right: 18px;
    font-size: 0.68rem;
    color: #1B3A6B;
    opacity: 0.25;
    letter-spacing: 0.02em;
    pointer-events: none;
}

.page-header {
    background: #1B3A6B;
    border-radius: 14px;
    padding: 20px 28px;
    display: flex;
    align-items: center;
    margin-bottom: 24px;
}
.page-header-left { display: flex; align-items: center; gap: 14px; }
.page-accent { width: 4px; height: 38px; background: #E87722; border-radius: 4px; }
.page-title { color: #fff; font-size: 1.1rem; font-weight: 600; margin: 0; }
.page-sub { color: #7A9BC4; font-size: 0.8rem; margin-top: 3px; }

.step-box {
    background: #fff;
    border: 1px solid #E8ECF2;
    border-radius: 14px;
    padding: 22px 26px;
    margin-bottom: 14px;
}
.step-label {
    font-size: 0.68rem;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 0.09em;
    color: #E87722;
    margin-bottom: 4px;
}
.step-title { font-size: 0.95rem; font-weight: 600; color: #1B3A6B; margin-bottom: 4px; }
.step-desc { font-size: 0.8rem; color: #8A96A8; line-height: 1.5; margin-bottom: 0; }

.alert-error {
    background: #FFF0F0;
    border: 1px solid #FFCDD2;
    border-left: 4px solid #E53935;
    border-radius: 8px;
    padding: 12px 16px;
    font-size: 0.82rem;
    color: #C62828;
    margin: 12px 0;
}
.alert-success {
    background: #F0FFF4;
    border: 1px solid #C8E6C9;
    border-left: 4px solid #2D9E5F;
    border-radius: 8px;
    padding: 12px 16px;
    font-size: 0.82rem;
    color: #1B5E20;
    margin: 12px 0;
}
.alert-info {
    background: #EAF2FF;
    border: 1px solid #BBDEFB;
    border-left: 4px solid #1B3A6B;
    border-radius: 8px;
    padding: 12px 16px;
    font-size: 0.82rem;
    color: #1B3A6B;
    margin: 12px 0;
}

.stat-row {
    display: flex;
    gap: 12px;
    margin: 16px 0;
}
.stat-card {
    flex: 1;
    background: #fff;
    border: 1px solid #E8ECF2;
    border-radius: 10px;
    padding: 14px;
    text-align: center;
}
.stat-val { font-size: 1.4rem; font-weight: 600; color: #1B3A6B; }
.stat-lbl { font-size: 0.72rem; color: #8A96A8; margin-top: 2px; }
.stat-card.rojo .stat-val { color: #E53935; }
.stat-card.verde .stat-val { color: #2D9E5F; }

[data-testid="stDownloadButton"] button {
    background-color: #2D9E5F !important;
    color: white !important;
    border: none !important;
    border-radius: 8px !important;
    font-size: 0.85rem !important;
    font-weight: 600 !important;
    width: 100% !important;
    padding: 10px 0 !important;
}
[data-testid="stDownloadButton"] button:hover { background-color: #1B3A6B !important; }
div[data-testid="stButton"] button {
    background: rgba(27,58,107,0.07) !important;
    color: #1B3A6B !important;
    border: 1px solid #D0D8E8 !important;
    border-radius: 8px !important;
    font-size: 0.8rem !important;
    font-weight: 500 !important;
}
div[data-testid="stButton"] button:hover { background: #1B3A6B !important; color: #fff !important; }
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="page-header">
    <div class="page-header-left">
        <div class="page-accent"></div>
        <div>
            <div class="page-title">📊 Control de Facturas Proveedores</div>
            <div class="page-sub">Compras · SOCOVAL</div>
        </div>
    </div>
</div>
""", unsafe_allow_html=True)

st.markdown('<div class="firma">Automatizaciones desarrolladas por Pedro Muñoz Ballier</div>', unsafe_allow_html=True)

if st.button("← Volver al panel"):
    st.switch_page("app.py")

st.markdown("""
<div class="step-box">
    <div class="step-label">Paso 1</div>
    <div class="step-title">Sube el archivo exportado desde SAP</div>
    <div class="step-desc">Formatos aceptados: .csv (separador ; · Latin-1) o .xlsx exportado desde SAP Business One.</div>
</div>
""", unsafe_allow_html=True)

archivo = st.file_uploader("", type=["csv", "xlsx"], label_visibility="collapsed")

if archivo is not None:
    try:
        if archivo.name.endswith('.xlsx'):
            df = pd.read_excel(archivo, engine='openpyxl')
        else:
            df = pd.read_csv(
                archivo,
                sep=';',
                encoding='latin-1',
                decimal=',',
                thousands='.'
            )

        if df.empty:
            st.markdown('<div class="alert-error">El archivo esta vacio. Sube un archivo con datos.</div>', unsafe_allow_html=True)
            st.stop()

        if len(df.columns) not in [13, 14]:
            st.markdown(f'<div class="alert-error">El archivo no tiene el formato correcto. Se esperaban 13 o 14 columnas y se encontraron {len(df.columns)}. Verifica que sea el export correcto de SAP.</div>', unsafe_allow_html=True)
            st.stop()

        # ── Renombrar columnas ────────────────────────────────
        if len(df.columns) == 14:
            df.columns = [
                'ID', 'Nro_Factura', 'Cod_Proveedor', 'RUT', 'Proveedor',
                'Fecha_Contabilizacion', 'Fecha_Vencimiento', 'Nro_Primario',
                'Cod_Proyecto', 'Nombre_Proyecto', 'Neto', 'IVA', 'Bruto',
                'Eliminar'
            ]
            df = df.drop(columns=['Eliminar', 'Neto', 'IVA'])
        else:
            df.columns = [
                'ID', 'Nro_Factura', 'Cod_Proveedor', 'RUT', 'Proveedor',
                'Fecha_Contabilizacion', 'Fecha_Vencimiento', 'Nro_Primario',
                'Cod_Proyecto', 'Nombre_Proyecto', 'Neto', 'IVA', 'Bruto'
            ]
            df = df.drop(columns=['Neto', 'IVA'])

        # ── Convertir fechas ──────────────────────────────────
        df['Fecha_Vencimiento']     = pd.to_datetime(df['Fecha_Vencimiento'],     dayfirst=True, errors='coerce')
        df['Fecha_Contabilizacion'] = pd.to_datetime(df['Fecha_Contabilizacion'], dayfirst=True, errors='coerce')

        # ── Eliminar duplicados ───────────────────────────────
        antes = len(df)
        df = df.drop_duplicates(subset=['Nro_Factura', 'Cod_Proveedor', 'Bruto']).copy()
        duplicados_eliminados = antes - len(df)

        hoy = pd.Timestamp.today().normalize()
        df['Dias'] = (df['Fecha_Vencimiento'] - hoy).dt.days

        # ── Vencidas ──────────────────────────────────────────
        vencidas = df[df['Dias'] < 0].copy()
        vencidas['Dias_Vencido'] = vencidas['Dias'].abs()
        vencidas['0-30 dias']  = vencidas['Bruto'].where(vencidas['Dias_Vencido'] <= 30)
        vencidas['30-60 dias'] = vencidas['Bruto'].where((vencidas['Dias_Vencido'] > 30) & (vencidas['Dias_Vencido'] <= 60))
        vencidas['60-90 dias'] = vencidas['Bruto'].where((vencidas['Dias_Vencido'] > 60) & (vencidas['Dias_Vencido'] <= 90))
        vencidas['Mas_90 dias'] = vencidas['Bruto'].where(vencidas['Dias_Vencido'] > 90)
        vencidas = vencidas.sort_values('Dias_Vencido', ascending=True)

        # ── Por vencer ────────────────────────────────────────
        por_vencer = df[df['Dias'] >= 0].copy()
        por_vencer['Dias_Faltantes'] = por_vencer['Dias']
        por_vencer['0-30 dias']  = por_vencer['Bruto'].where(por_vencer['Dias_Faltantes'] <= 30)
        por_vencer['30-60 dias'] = por_vencer['Bruto'].where((por_vencer['Dias_Faltantes'] > 30) & (por_vencer['Dias_Faltantes'] <= 60))
        por_vencer['60-90 dias'] = por_vencer['Bruto'].where((por_vencer['Dias_Faltantes'] > 60) & (por_vencer['Dias_Faltantes'] <= 90))
        por_vencer = por_vencer.sort_values('Dias_Faltantes', ascending=True)

        st.markdown(f"""
        <div class="alert-success">
            Archivo cargado correctamente — {len(df)} facturas unicas · {duplicados_eliminados} duplicados eliminados
        </div>
        """, unsafe_allow_html=True)

        st.markdown("""
        <div class="step-box">
            <div class="step-label">Paso 2</div>
            <div class="step-title">Resumen del analisis</div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown(f"""
        <div class="stat-row">
            <div class="stat-card rojo">
                <div class="stat-val">{len(vencidas)}</div>
                <div class="stat-lbl">Facturas vencidas</div>
            </div>
            <div class="stat-card verde">
                <div class="stat-val">{len(por_vencer)}</div>
                <div class="stat-lbl">Por vencer</div>
            </div>
            <div class="stat-card">
                <div class="stat-val">{len(df)}</div>
                <div class="stat-lbl">Total facturas</div>
            </div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("""
        <div class="step-box" style="margin-top:16px;">
            <div class="step-label">Paso 3</div>
            <div class="step-title">Vista previa — Facturas Vencidas</div>
        </div>
        """, unsafe_allow_html=True)
        st.dataframe(vencidas.drop(columns=['Dias']), use_container_width=True)

        st.markdown("""
        <div class="step-box" style="margin-top:16px;">
            <div class="step-label">Paso 4</div>
            <div class="step-title">Vista previa — Facturas Por Vencer</div>
        </div>
        """, unsafe_allow_html=True)
        st.dataframe(por_vencer.drop(columns=['Dias']), use_container_width=True)

        # ── Generar Excel ─────────────────────────────────────
        fecha_str = hoy.strftime('%d-%m-%Y')
        nombre_archivo = f'Resultado control de Facturas - {fecha_str}.xlsx'
        FIRMA = 'Automatizacion creada por Pedro Munoz Ballier'

        # ── Construir hoja "Todas las Facturas" ───────────────
        vencidas_todas   = vencidas.drop(columns=['Dias']).copy()
        por_vencer_todas = por_vencer.drop(columns=['Dias']).copy()

        vencidas_todas['Estado']   = 'Vencida'
        por_vencer_todas['Estado'] = 'Por Vencer'

        # Renombrar columnas de dias para unificar
        vencidas_todas   = vencidas_todas.rename(columns={'Dias_Vencido': 'Dias'})
        vencidas_todas['Dias'] = vencidas_todas['Dias'] * -1
        por_vencer_todas = por_vencer_todas.rename(columns={'Dias_Faltantes': 'Dias'})

        # Insertar columna Estado entre Nombre_Proyecto y Bruto
        cols_vencidas   = list(vencidas_todas.columns)
        cols_por_vencer = list(por_vencer_todas.columns)

        def reordenar(df):
            cols = [c for c in df.columns if c not in ['Estado', 'Bruto', 'Dias']]
            idx = cols.index('Nombre_Proyecto') + 1
            cols.insert(idx, 'Estado')
            cols.append('Bruto')
            cols.append('Dias')
            tramos = [c for c in df.columns if c not in cols]
            return df[cols + tramos]

        vencidas_todas   = reordenar(vencidas_todas)
        por_vencer_todas = reordenar(por_vencer_todas)

        todas = pd.concat([vencidas_todas, por_vencer_todas], ignore_index=True)

        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            todas.to_excel(writer,                          sheet_name='Todas las Facturas', index=False)
            vencidas.drop(columns=['Dias']).to_excel(writer,   sheet_name='Vencidas',        index=False)
            por_vencer.drop(columns=['Dias']).to_excel(writer, sheet_name='Por Vencer',      index=False)

        buffer.seek(0)
        wb = load_workbook(buffer)

        for nombre_hoja in ['Todas las Facturas', 'Vencidas', 'Por Vencer']:
            ws = wb[nombre_hoja]
            col_fin  = get_column_letter(ws.max_column)
            fila_fin = ws.max_row
            rango    = f'A1:{col_fin}{fila_fin}'
            tabla    = Table(displayName=nombre_hoja.replace(' ', '_'), ref=rango)
            estilo   = TableStyleInfo(
                name='TableStyleMedium9',
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            tabla.tableStyleInfo = estilo
            ws.add_table(tabla)

            fila_firma = fila_fin + 5
            col_firma  = ws.max_column
            celda      = ws.cell(row=fila_firma, column=col_firma, value=FIRMA)
            celda.font = Font(italic=True, color='808080', size=9)
            celda.alignment = Alignment(horizontal='right')

        buffer_final = BytesIO()
        wb.save(buffer_final)
        buffer_final.seek(0)

        st.markdown("<div style='margin-top:16px;'>", unsafe_allow_html=True)
        st.download_button(
            label="⬇️ Descargar Excel — Control de Facturas",
            data=buffer_final,
            file_name=nombre_archivo,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        st.markdown("</div>", unsafe_allow_html=True)

    except Exception as e:
        st.markdown(f'<div class="alert-error">Error al procesar el archivo: {str(e)}</div>', unsafe_allow_html=True)
